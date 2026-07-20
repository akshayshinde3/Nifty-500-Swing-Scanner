import io
import datetime as dt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_prebreakout_excel(daily_watch, weekly_watch, confirmed_daily=None, confirmed_weekly=None):
    """Generate a professionally styled Excel workbook containing pre-breakout stock candidates.
    Returns an openpyxl Workbook object.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    sub_font = Font(name="Calibri", size=10, italic=True, color="595959")
    sub_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    cell_font = Font(name="Calibri", size=11)
    bold_cell_font = Font(name="Calibri", size=11, bold=True)
    
    # Accent fills
    buy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # light green
    sl_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # light red
    target_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # light blue
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")     # light grey row

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "Symbol",
        "Timeframe",
        "Pattern / Setup",
        "LTP (₹)",
        "Buy Trigger (₹)",
        "Stop Loss (₹)",
        "Target 1 (₹)",
        "Target 2 (₹)",
        "Risk:Reward",
        "Confidence Score",
        "Expected Breakout",
        "Analysis / Reasons"
    ]

    def populate_sheet(ws, title_text, items, is_combined=False):
        ws.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws.merge_cells("A1:L1")
        ws["A1"] = f"📈 NIFTY 500 — {title_text.upper()}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 36

        # Subtitle Block
        ws.merge_cells("A2:L2")
        now_str = dt.datetime.now().strftime("%d-%b-%Y %I:%M %p IST")
        ws["A2"] = f"Generated on {now_str} | Total Pre-Breakout Candidates: {len(items)} | Confidential Watchlist"
        ws["A2"].font = sub_font
        ws["A2"].fill = sub_fill
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20

        # Header Row
        ws.row_dimensions[4].height = 28
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if col_idx not in (3, 12) else "left", vertical="center", wrap_text=True)

        if not items:
            ws.merge_cells("A5:L5")
            cell = ws.cell(row=5, column=1, value="No pre-breakout candidates detected at this time.")
            cell.font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[5].height = 30
            return

        # Data Rows
        row_idx = 5
        for item in items:
            ws.row_dimensions[row_idx].height = 24
            is_alt = (row_idx % 2 == 0)
            row_fill = alt_fill if is_alt else None

            reasons_str = "; ".join(item.get("reasons", [])) if isinstance(item.get("reasons"), list) else str(item.get("reasons", ""))
            tf = item.get("timeframe", "Daily")

            values = [
                (item.get("symbol", ""), bold_cell_font, Alignment(horizontal="center", vertical="center"), None, None),
                (tf, cell_font, Alignment(horizontal="center", vertical="center"), None, None),
                (item.get("strategy_used", ""), cell_font, Alignment(horizontal="left", vertical="center"), None, None),
                (item.get("ltp", 0.0), cell_font, Alignment(horizontal="right", vertical="center"), '₹ #,##0.00', None),
                (item.get("buy_price", 0.0), bold_cell_font, Alignment(horizontal="right", vertical="center"), '₹ #,##0.00', buy_fill),
                (item.get("stop_loss", 0.0), cell_font, Alignment(horizontal="right", vertical="center"), '₹ #,##0.00', sl_fill),
                (item.get("target1", 0.0), cell_font, Alignment(horizontal="right", vertical="center"), '₹ #,##0.00', buy_fill),
                (item.get("target2", 0.0), cell_font, Alignment(horizontal="right", vertical="center"), '₹ #,##0.00', target_fill),
                (item.get("risk_reward", ""), cell_font, Alignment(horizontal="center", vertical="center"), None, None),
                (item.get("confidence", 0) / 100.0, bold_cell_font, Alignment(horizontal="center", vertical="center"), '0%', None),
                (item.get("expected_breakout_date", ""), cell_font, Alignment(horizontal="center", vertical="center"), None, None),
                (reasons_str, cell_font, Alignment(horizontal="left", vertical="center"), None, None)
            ]

            for c_idx, (val, fnt, align, num_fmt, custom_fill) in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=c_idx, value=val)
                c.font = fnt
                c.alignment = align
                c.border = thin_border
                if custom_fill:
                    c.fill = custom_fill
                elif row_fill:
                    c.fill = row_fill

                if num_fmt:
                    c.number_format = num_fmt

            row_idx += 1

        # Auto-fit column widths
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.row in (1, 2):  # skip merged title rows
                    continue
                if cell.value is not None:
                    val_str = str(cell.value)
                    if cell.number_format and ('₹' in cell.number_format or '%' in cell.number_format):
                        val_str = f"₹ {val_str}"
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        ws.column_dimensions['L'].width = 50  # Reasons column wider

    # 1. Combined Sheet
    combined_items = []
    for d in daily_watch:
        item = dict(d)
        item["timeframe"] = "Daily"
        combined_items.append(item)
    for w in weekly_watch:
        item = dict(w)
        item["timeframe"] = "Weekly"
        combined_items.append(item)

    combined_items.sort(key=lambda x: -x.get("confidence", 0))

    ws_all = wb.create_sheet(title="All Pre-Breakouts")
    populate_sheet(ws_all, "All Pre-Breakout Watchlist", combined_items, is_combined=True)

    # 2. Daily Pre-Breakouts Sheet
    ws_daily = wb.create_sheet(title="Daily Pre-Breakouts")
    populate_sheet(ws_daily, "Daily Pre-Breakout Watchlist", [dict(x, timeframe="Daily") for x in daily_watch])

    # 3. Weekly Pre-Breakouts Sheet
    ws_weekly = wb.create_sheet(title="Weekly Pre-Breakouts")
    populate_sheet(ws_weekly, "Weekly Pre-Breakout Watchlist", [dict(x, timeframe="Weekly") for x in weekly_watch])

    # 4. Confirmed Breakouts Sheet (Optional reference tab)
    if confirmed_daily or confirmed_weekly:
        confirmed_items = []
        for d in (confirmed_daily or []):
            item = dict(d)
            item["timeframe"] = "Daily"
            confirmed_items.append(item)
        for w in (confirmed_weekly or []):
            item = dict(w)
            item["timeframe"] = "Weekly"
            confirmed_items.append(item)
        confirmed_items.sort(key=lambda x: -x.get("confidence", 0))

        ws_conf = wb.create_sheet(title="Confirmed Breakouts")
        populate_sheet(ws_conf, "Confirmed Breakout Alerts", confirmed_items)

    return wb
